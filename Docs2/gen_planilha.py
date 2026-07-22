# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

VERDE = "1E6B52"
CINZA = "F2F2F2"
AMARELO = "FFF2CC"
Arial = "Arial"

wb = openpyxl.Workbook()

# ============ LEIA-ME ============
ws = wb.active
ws.title = "LEIA-ME"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 95

rows = [
    ("T", "PLANILHA DE EQUIPAMENTOS LOCADOS — MODELO SANEADO v1.0", ""),
    ("S", "Base de dados do Agente de Locações (demo). Julho/2026.", ""),
    ("", "", ""),
    ("H", "Como usar", ""),
    ("I", "1.", "Preencha apenas as colunas com cabeçalho AMARELO na aba CONTRATOS. As colunas cinza são calculadas automaticamente — não digite nelas."),
    ("I", "2.", "Campos marcados com * são OBRIGATÓRIOS. A coluna CADASTRO mostra 'FALTA: ...' enquanto algo obrigatório estiver vazio. O agente NÃO monitora contratos com cadastro incompleto — ele avisa a pendência."),
    ("I", "3.", "DATA INÍCIO é a data-base da cobrança (1º dia do contrato). O PRÓXIMO VENCIMENTO é calculado a partir dela conforme o CICLO."),
    ("I", "4.", "CICLO = DATA ESPECÍFICA exige preencher DATA FIM (ex.: locação de 1 dia para concretagem)."),
    ("I", "5.", "SITUAÇÃO é atualizada pelo agente quando a Daniela responde no WhatsApp — mas pode ser alterada manualmente a qualquer momento."),
    ("I", "6.", "A aba LOG é preenchida somente pelo agente (histórico de alertas e respostas). Não edite."),
    ("I", "7.", "As listas suspensas (obra, fornecedor, ciclo, situação) vêm da aba LISTAS — inclua novos fornecedores/obras lá."),
    ("", "", ""),
    ("H", "Significado da coluna STATUS ALERTA", ""),
    ("I", "OK", "Vencimento a mais de 5 dias. Nenhuma ação."),
    ("I", "ALERTAR", "Vence em 5 dias ou menos. O agente envia alerta no WhatsApp."),
    ("I", "VENCIDO", "Ciclo virou sem decisão registrada. Alerta com prioridade."),
    ("I", "DEVOLVER", "Devolução solicitada e ainda não confirmada. O agente cobra a confirmação."),
    ("", "", ""),
    ("H", "Regras de datas", ""),
    ("I", "Fim de semana", "Vencimento em sábado/domingo/segunda gera alerta já na quinta-feira (evita pagar o fim de semana)."),
    ("I", "Fonte dos dados", "Linhas de exemplo extraídas da planilha real de locações (obras ARBO e LOT. CELEBRATION), com datas ajustadas para a demo. Valores conforme planilha original — validar com a Daniela."),
]
r = 1
for kind, b, c in rows:
    if kind == "T":
        ws.cell(row=r, column=2, value=b).font = Font(name=Arial, bold=True, size=14, color=VERDE)
    elif kind == "S":
        ws.cell(row=r, column=2, value=b).font = Font(name=Arial, size=10, color="595959")
    elif kind == "H":
        ws.cell(row=r, column=2, value=b).font = Font(name=Arial, bold=True, size=11)
    elif kind == "I":
        ws.cell(row=r, column=2, value=b).font = Font(name=Arial, bold=True, size=10)
        cc = ws.cell(row=r, column=3, value=c)
        cc.font = Font(name=Arial, size=10)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ============ LISTAS ============
wl = wb.create_sheet("LISTAS")
listas = {
    "A": ("OBRAS", ["ARBO", "PARADISO", "LOTEAMENTO CELEBRATION"]),
    "B": ("CICLOS", ["MENSAL", "QUINZENAL", "SEMANAL", "DIÁRIA", "DATA ESPECÍFICA"]),
    "C": ("SITUAÇÕES", ["EM OBRA", "DEVOLUÇÃO SOLICITADA", "DEVOLVIDO", "RENOVADO"]),
    "D": ("FORNECEDORES", ["LOCAFAZ", "ORIGINAL", "LOCSOLO", "LOCARBEM", "PAMPULHA ANDAIMES", "ATEX", "TUPI ANDAIMES", "MINAS LOCC"]),
}
for col, (titulo, itens) in listas.items():
    wl[f"{col}1"] = titulo
    wl[f"{col}1"].font = Font(name=Arial, bold=True, size=10, color="FFFFFF")
    wl[f"{col}1"].fill = PatternFill("solid", fgColor=VERDE)
    for i, it in enumerate(itens, start=2):
        wl[f"{col}{i}"] = it
        wl[f"{col}{i}"].font = Font(name=Arial, size=10)
    wl.column_dimensions[col].width = 26

# ============ CONTRATOS ============
wc = wb.create_sheet("CONTRATOS")
headers = [
    ("ID", 6, "auto"),          # A
    ("OBRA *", 22, "in"),        # B
    ("EQUIPAMENTO / MATERIAL *", 34, "in"),  # C
    ("QUANT. *", 8, "in"),       # D
    ("FORNECEDOR *", 20, "in"),  # E
    ("Nº CONTRATO *", 13, "in"), # F
    ("VALOR DO CICLO (R$) *", 14, "in"),  # G
    ("DATA INÍCIO *", 13, "in"), # H
    ("CICLO *", 15, "in"),       # I
    ("DATA FIM (se ciclo = data específica)", 14, "in"),  # J
    ("RESPONSÁVEL NA OBRA", 20, "in"),  # K
    ("TELEFONE RESPONSÁVEL", 16, "in"), # L
    ("SITUAÇÃO *", 22, "in"),    # M
    ("PRÓXIMO VENCIMENTO", 14, "calc"),  # N
    ("DIAS P/ VENCER", 10, "calc"),      # O
    ("STATUS ALERTA", 12, "calc"),       # P
    ("CADASTRO", 26, "calc"),            # Q
    ("ÚLTIMA RESPOSTA (agente)", 26, "agent"),  # R
    ("DATA RESPOSTA (agente)", 13, "agent"),    # S
    ("OBSERVAÇÕES", 30, "in"),   # T
]
thin = Border(*[Side(style="thin", color="BFBFBF")]*4)
for i, (h, w, kind) in enumerate(headers, start=1):
    c = wc.cell(row=1, column=i, value=h)
    fill = AMARELO if kind == "in" else (CINZA if kind in ("calc", "agent") else "FFFFFF")
    color = "333333"
    if kind == "in":
        c.fill = PatternFill("solid", fgColor="B78A00" if False else "FFD966")
    else:
        c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.font = Font(name=Arial, bold=True, size=9, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = thin
    wc.column_dimensions[get_column_letter(i)].width = w
wc.row_dimensions[1].height = 42
wc.freeze_panes = "D2"

# dados de exemplo (da planilha real, datas ajustadas p/ demo em 15/07/2026)
dados = [
    # OBRA, EQUIP, QT, FORN, NCONTR, VALOR, INICIO, CICLO, DATAFIM, RESP, TEL, SIT, ULTRESP, DTRESP, OBS
    ("ARBO", "MARTELO DEMOLIDOR TE2000 - M20", 1, "LOCAFAZ", "CS003087", 900.00, date(2026,4,13), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, ""),
    ("ARBO", "BOMBA SUBMERSÍVEL 3/4'' SAPO", 1, "ORIGINAL", "2253", 220.00, date(2026,5,7), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, ""),
    ("ARBO", "MINI GRUA 500KG + PAINEL NR-12", 1, "ORIGINAL", "1645", 915.00, date(2026,4,8), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, "Grua 825 + painel 90"),
    ("ARBO", "PERFURATRIZ DE COLUNA DD 200", 1, "ORIGINAL", "1339", 2900.00, date(2026,4,14), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, ""),
    ("ARBO", "TORRE COMPLETA", 1, "LOCSOLO", "218158", 1298.50, date(2026,3,26), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, ""),
    ("ARBO", "150 FORCADO SIMPLES", 1, "LOCSOLO", "220366", 600.00, date(2026,6,17), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "DEVOLUÇÃO SOLICITADA", "Daniela: pode devolver", date(2026,7,10), "Aguardando retirada"),
    ("ARBO", "250 ESCORAS METÁLICAS 3,10M", 1, "PAMPULHA ANDAIMES", "6769", 2922.00, date(2026,4,23), "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, ""),
    ("LOTEAMENTO CELEBRATION", "BETONEIRA À GASOLINA 400L", 1, "LOCAFAZ", "3936", 650.00, date(2026,4,30), "MENSAL", None, "Almoxarife CELEBRATION", "(31) 9xxxx-0002", "EM OBRA", "", None, ""),
    ("LOTEAMENTO CELEBRATION", "CONTAINER BANHEIRO (302)", 1, "MINAS LOCC", "CT-3367", 280.00, date(2026,3,24), "MENSAL", None, "Almoxarife CELEBRATION", "(31) 9xxxx-0002", "EM OBRA", "", None, ""),
    ("ARBO", "MOTOR VIBRADOR ALTA FREQUÊNCIA", 1, "LOCAFAZ", "CS001218", 350.00, date(2026,7,17), "DATA ESPECÍFICA", date(2026,7,18), "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, "Concretagem de sábado — devolver segunda cedo"),
    # exemplos de cadastro INCOMPLETO (de propósito):
    ("ARBO", "MOTOR VIBRADOR ALTA FREQUÊNCIA", 1, "LOCAFAZ", "CW182296", 200.00, None, "MENSAL", None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, "EXEMPLO: falta DATA INÍCIO — corrigir com a NF"),
    ("ARBO", "30M DE CONDUTOR DE ENTULHO", 1, "LOCSOLO", "219793", 1350.00, date(2026,6,9), None, None, "Almoxarife ARBO", "(31) 9xxxx-0001", "EM OBRA", "", None, "EXEMPLO: falta CICLO (data original digitada errada: '3117')"),
]

nrows = len(dados)
for idx, d in enumerate(dados, start=2):
    (obra, eq, qt, forn, nc, val, ini, ciclo, dfim, resp, tel, sit, ultr, dtr, obs) = d
    wc.cell(row=idx, column=1, value=idx-1)
    vals = {2: obra, 3: eq, 4: qt, 5: forn, 6: nc, 7: val, 8: ini, 9: ciclo, 10: dfim,
            11: resp, 12: tel, 13: sit, 18: ultr, 19: dtr, 20: obs}
    for col, v in vals.items():
        if v is not None and v != "":
            wc.cell(row=idx, column=col, value=v)
    r = idx
    # N: próximo vencimento
    wc.cell(row=r, column=14, value=(
        f'=IF(OR($H{r}="",$I{r}="",$M{r}="DEVOLVIDO"),"",'
        f'IF($I{r}="DATA ESPECÍFICA",IF($J{r}="","",$J{r}),'
        f'IF($I{r}="MENSAL",EDATE($H{r},DATEDIF($H{r},TODAY(),"m")+1),'
        f'IF($I{r}="QUINZENAL",$H{r}+(INT((TODAY()-$H{r})/14)+1)*14,'
        f'IF($I{r}="SEMANAL",$H{r}+(INT((TODAY()-$H{r})/7)+1)*7,'
        f'IF($I{r}="DIÁRIA",TODAY()+1,""))))))'))
    # O: dias p/ vencer
    wc.cell(row=r, column=15, value=f'=IF($N{r}="","",$N{r}-TODAY())')
    # P: status alerta
    wc.cell(row=r, column=16, value=(
        f'=IF($M{r}="DEVOLVIDO","—",IF($M{r}="DEVOLUÇÃO SOLICITADA","DEVOLVER",'
        f'IF($N{r}="","",IF($O{r}<0,"VENCIDO",IF($O{r}<=5,"ALERTAR","OK")))))'))
    # Q: cadastro
    tj = (f'_xlfn.TEXTJOIN(", ",TRUE,'
          f'IF($B{r}="","OBRA",""),IF($C{r}="","EQUIPAMENTO",""),IF($E{r}="","FORNECEDOR",""),'
          f'IF($F{r}="","Nº CONTRATO",""),IF($G{r}="","VALOR",""),IF($H{r}="","DATA INÍCIO",""),'
          f'IF($I{r}="","CICLO",""),IF(AND($I{r}="DATA ESPECÍFICA",$J{r}=""),"DATA FIM",""))')
    wc.cell(row=r, column=17, value=f'=IF({tj}="","OK","FALTA: "&{tj})')

# estilos das linhas
fmt_date = "DD/MM/YYYY"
fmt_money = '#,##0.00'
for r in range(2, nrows+2):
    for c in range(1, 21):
        cell = wc.cell(row=r, column=c)
        cell.font = Font(name=Arial, size=10)
        cell.border = thin
        if c in (8, 10, 14, 19):
            cell.number_format = fmt_date
        if c == 7:
            cell.number_format = fmt_money
        if c in (15,):
            cell.number_format = "0"
    if r % 2 == 0:
        pass

# validações
dv_obra = DataValidation(type="list", formula1="=LISTAS!$A$2:$A$20", allow_blank=True)
dv_ciclo = DataValidation(type="list", formula1="=LISTAS!$B$2:$B$6", allow_blank=True)
dv_sit = DataValidation(type="list", formula1="=LISTAS!$C$2:$C$5", allow_blank=True)
dv_forn = DataValidation(type="list", formula1="=LISTAS!$D$2:$D$40", allow_blank=True)
for dv, colrange in ((dv_obra, "B2:B200"), (dv_ciclo, "I2:I200"), (dv_sit, "M2:M200"), (dv_forn, "E2:E200")):
    wc.add_data_validation(dv)
    dv.add(colrange)

# formatação condicional STATUS ALERTA
rng = "P2:P200"
wc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"VENCIDO"'], fill=PatternFill("solid", fgColor="F4CCCC"), font=Font(color="990000", bold=True)))
wc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"ALERTAR"'], fill=PatternFill("solid", fgColor="FFF2CC"), font=Font(color="7F6000", bold=True)))
wc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"DEVOLVER"'], fill=PatternFill("solid", fgColor="FCE5CD"), font=Font(color="B45F06", bold=True)))
wc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor="D9EAD3"), font=Font(color="274E13")))
# CADASTRO incompleto
wc.conditional_formatting.add("Q2:Q200", FormulaRule(formula=['LEFT($Q2,5)="FALTA"'], fill=PatternFill("solid", fgColor="F4CCCC"), font=Font(color="990000", bold=True)))

# ============ LOG ============
wg = wb.create_sheet("LOG")
log_headers = ["DATA/HORA", "ID CONTRATO", "Nº CONTRATO", "EVENTO", "DETALHE", "CANAL", "AUTOR"]
for i, h in enumerate(log_headers, start=1):
    c = wg.cell(row=1, column=i, value=h)
    c.font = Font(name=Arial, bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE)
    wg.column_dimensions[get_column_letter(i)].width = [18, 11, 13, 26, 55, 12, 16][i-1]
# exemplo ilustrativo
exemplo_log = [
    ("15/07/2026 07:00", 5, "218158", "ALERTA ENVIADO", "TORRE COMPLETA (LOCSOLO) renova em 26/07 — R$ 1.298,50. Devolver ou renovar?", "WhatsApp", "Agente"),
    ("15/07/2026 08:12", 5, "218158", "RESPOSTA RECEBIDA", "“pode devolver, obra já desmontou”", "WhatsApp", "Daniela"),
    ("15/07/2026 08:12", 5, "218158", "STATUS ALTERADO", "EM OBRA → DEVOLUÇÃO SOLICITADA", "Sistema", "Agente"),
]
for ri, row in enumerate(exemplo_log, start=2):
    for ci, v in enumerate(row, start=1):
        c = wg.cell(row=ri, column=ci, value=v)
        c.font = Font(name=Arial, size=9, italic=True, color="808080")
wg.cell(row=5, column=1, value="(linhas acima são exemplo ilustrativo — o agente preenche esta aba automaticamente)").font = Font(name=Arial, size=9, italic=True, color="808080")

wb.save("Planilha Locações - Modelo Saneado v1.0.xlsx")
print("OK")
